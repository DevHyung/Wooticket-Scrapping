from openpyxl import Workbook
from openpyxl import load_workbook
import time
import os
FILENAME = 'sample.xlsx'
if os.path.isfile(FILENAME): # 파일있는 경우
    new_row = ['data1', 'data2', 'data3', 'data4']
    wb = load_workbook(filename=FILENAME)
    ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
    row = ws.max_row + 3
    ws.cell(row=row, column=2).value = new_row[0]
    ws.append(new_row)
    wb.save(FILENAME)
else:
    now = time.localtime()
    s = "%04d-%02d-%02d %02d:%02d:%02d" % \
        (now.tm_year, now.tm_mon, now.tm_mday, now.tm_hour, now.tm_min, now.tm_sec)

    book = Workbook()
    sheet = book.active
    sheet.title = '테스트'

    header = ['일자','긁어온 시간','']
    sheet.append(header)

    sheet.cell(row=2, column=2).value = s

    now = time.strftime("%x")
    sheet['A2'] = now

    rows = [
        [88, 46, 57],
        [89, 38, 12],
        [23, 59, 78],
        [56, 21, 98],
        [24, 18, 43],
        [34, 15, 67]
    ]

    for row in rows:
        sheet.append(row)

    print(now)
    book.save(FILENAME)

"""



"""