#-*-encoding:utf8-*-
from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook
from openpyxl import load_workbook
import time
import os

def valid_user():
    # 20180730 10:03기준 6시간
    now = 1533091486.2744226
    terminTime = now + 60 * 60 * 6
    print("체험판 만료기간 : ", time.ctime(terminTime))
    if time.time() > terminTime:
        print('만료되었습니다.')
        exit(-1)
    else:
        print(">>> 프로그램이 실행되었습니다.")


def auto_fit_width():
    dims = {}
    for row in sheet1.rows:
        for cell in row:
            if cell.value:
                dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))
    for col, value in dims.items():
        sheet1.column_dimensions[col].width = value

def get_bs_obejct_by_url(url):
    html = requests.get(url)
    # print(html.encoding) # ISO-8859-1 인코딩나와서
    html.encoding = 'euc-kr'  # 한글 인코딩으로 변환
    return BeautifulSoup(html.text, 'lxml')

def remove_dup_data_at_list(tmp):
    returnList = []
    returnList.append(tmp[0])
    for idx in range(1,len(tmp)):
        if tmp[idx-1] != tmp[idx]:
            returnList.append(tmp[idx])
        else:
            returnList.append('')
    return returnList
if __name__=="__main__":
    #valid_user()
    #===    CONFIG
    #FILENAME = r'C:\Users\khuph\Desktop\DATA.xlsx'
    f = open("CONFIG.txt",encoding='utf8').readlines()
    targetDir = f[0].split(':',maxsplit=1)[1].strip()
    targetName = f[1].split(':')[1].strip()
    FILENAME = targetDir+'\\'+targetName
    print(FILENAME)
    #===    DECLARE & DEFINE
    keyList = ['롯데백화점@50만원권', '롯데백화점@10만원권', '롯데백화점@5만원권', '롯데백화점@1만원권', '롯데@스페셜카드 상품권', '롯데@JTB 상품권 (종이권)',
               '롯데호텔@10만원권(종이식)', '신세계백화점@50만원권', '신세계백화점@10만원권', '신세계백화점@5만원권', '신세계백화점@1만원권', '현대백화점@50만원',
               '현대백화점@10만원권', '홈플러스@30만원권', '홈플러스@10만원권', '홈플러스@5만원권', 'AK플라자@10만원권', '삼성상품권@10만원권 (종이식)',
               '이랜드@상품권 10만원', '-@갤러리아백화점10만원', '메가마트@상품권', '아이파크@상품권 10만원권', '금강제화@상품권 10만원권', '금강제화@상품권 7만원권',
               '금강제화@상품권 5만원권', '삼성@기프트카드', '신한@기프트 카드', '현대@기프트 카드', '국민@기프트 카드', 'KEB하나@기프트카드', 'BC@기프트 카드',
               '롯데@기프트 카드', '각종@기프트카드 30만원권', '각종@기프트카드 10만원권', '각종@기프트카드 5만원권', '농협@기프트 카드 50만원권', '농협@기프트 카드 30만원권',
               '농협@기프트 카드 10만원권', '농산물(농협)상품권@10만원', '농산물(농협)@상품권 1만원권', '해피머니@문화상품권 1만원권', '문화상품권@1만원권', '문화상품권@5만원권',
               '쇼핑문화@5만원권', '도서문화상품권@1만원권', '도서문화상품권@5천원권', 'CJ@상품권 10만원', 'CJ@상품권 1만원권', 'CJ@외식상품권 5만원권',
               '국민@관광상품권10만원권', 'GS@주유권 5만원권', 'GS주유권@1만원권', 'SK주유권@1만원권', 'SK주유권@5만원', 'S-oil주유권@5만원권', '현대주유권@5만원',
               '하나투어@여행상품권', '모두투어@여행상품권', '토다이@상품권', '-@제일모직상품권10만원권(면세점가능)', 'LG패션@상품권 10만원권']
    bs4 = get_bs_obejct_by_url('http://www.wooticket.com/popup_price.php')
    headList = ['롯데백화점', '롯데백화점', '롯데백화점', '롯데백화점', '롯데', '롯데', '롯데호텔', '신세계백화점', '신세계백화점', '신세계백화점', '신세계백화점', '현대백화점',
                '현대백화점', '홈플러스', '홈플러스', '홈플러스', 'AK플라자', '삼성상품권', '이랜드', '-', '메가마트', '아이파크', '금강제화', '금강제화', '금강제화',
                '삼성', '신한', '현대', '국민', 'KEB하나', 'BC', '롯데', '각종', '각종', '각종', '농협', '농협', '농협', '농산물(농협)상품권',
                '농산물(농협)', '해피머니', '문화상품권', '문화상품권', '쇼핑문화', '도서문화상품권', '도서문화상품권', 'CJ', 'CJ', 'CJ', '국민', 'GS',
                'GS주유권', 'SK주유권', 'SK주유권', 'S-oil주유권', '현대주유권', '하나투어', '모두투어', '토다이', '-', 'LG패션']  # 맨위
    titleList = ['50만원권', '10만원권', '5만원권', '1만원권', '스페셜카드 상품권', 'JTB 상품권 (종이권)', '10만원권(종이식)', '50만원권', '10만원권', '5만원권',
                 '1만원권', '50만원', '10만원권', '30만원권', '10만원권', '5만원권', '10만원권', '10만원권 (종이식)', '상품권 10만원', '갤러리아백화점10만원',
                 '상품권', '상품권 10만원권', '상품권 10만원권', '상품권 7만원권', '상품권 5만원권', '기프트카드', '기프트 카드', '기프트 카드', '기프트 카드',
                 '기프트카드', '기프트 카드', '기프트 카드', '기프트카드 30만원권', '기프트카드 10만원권', '기프트카드 5만원권', '기프트 카드 50만원권',
                 '기프트 카드 30만원권', '기프트 카드 10만원권', '10만원', '상품권 1만원권', '문화상품권 1만원권', '1만원권', '5만원권', '5만원권', '1만원권',
                 '5천원권', '상품권 10만원', '상품권 1만원권', '외식상품권 5만원권', '관광상품권10만원권', '주유권 5만원권', '1만원권', '1만원권', '5만원', '5만원권',
                 '5만원', '여행상품권', '여행상품권', '상품권', '제일모직상품권10만원권(면세점가능)', '상품권 10만원권']  # 그아래
    buyList = ['' for i in range(61)]  # 매입
    buyPerList = ['' for i in range(61)]
    sellList = ['' for i in range(61)]  # 판매
    sellPerList = ['' for i in range(61)]
    spreadList = ['' for i in range(61)]
    now = time.localtime()
    nowDate = time.strftime("%x")
    nowTime = "%04d-%02d-%02d %02d:%02d:%02d" % \
        (now.tm_year, now.tm_mon, now.tm_mday, now.tm_hour, now.tm_min, now.tm_sec)

    header1 = ['일자', '긁어온 시간', '']
    header2 = [' ', ' ', ' ']
    datas = [nowDate, nowTime, ' ']


    #===    CODE
    #파싱
    trs = bs4.find_all('table')[3].find_all("tr")[1:]
    for tr in trs[1:-1]:
        try:
            name = tr.find_all('td')[1].get_text().strip()
            key = name.split(' ', maxsplit=1)[0]+'@'+name.split(' ', maxsplit=1)[1]
        except:
            key = '-@'+name.split(' ', maxsplit=1)[0]
        finally:
            keyIndex = keyList.index(key)
            buy = tr.find_all('td')[2].find('font').get_text().strip()
            buyPer = tr.find_all('td')[2].get_text().strip().split('(')[1].split(')')[0]
            buyList[keyIndex] = buy
            buyPerList[keyIndex] = buyPer

            sell = tr.find_all('td')[3].find('font').get_text().strip()
            sellPer = tr.find_all('td')[3].get_text().strip().split('(')[1].split(')')[0]
            sellPerList[keyIndex] = sellPer
            sellList[keyIndex] = sell
            spreadList[keyIndex] = int(sell.replace(',',''))-int(buy.replace(',',''))
    # 엑셀 저장 부분
    if os.path.isfile(FILENAME): # 파일있는 경우
        wb = load_workbook(filename=FILENAME)
        sheet1 = wb[wb.sheetnames[0]]
        #sheet1 = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        nextRow = sheet1.max_row + 1
        sheet1.append(datas + buyList + header2 + sellList)

        sheet2 = wb[wb.sheetnames[1]]
        #sheet2 = wb.get_sheet_by_name(wb.get_sheet_names()[1])
        nextRow = sheet2.max_row + 1
        sheet2.append(datas + buyPerList + header2 + sellPerList)

        sheet3 = wb[wb.sheetnames[2]]
        #sheet3 = wb.get_sheet_by_name(wb.get_sheet_names()[2])
        nextRow = sheet3.max_row + 1
        sheet3.append(datas + spreadList)
        wb.save(FILENAME)
    else: # 파일 없는 경우
        # 엑셀파일 초기설정
        book = Workbook()
        # 시트 설정
        sheet1 = book.active
        sheet1.column_dimensions['A'].width = 10
        sheet1.column_dimensions['B'].width = 20
        sheet1.column_dimensions['C'].width = 2
        sheet1.title = 'RawData'

        sheet2 = book.create_sheet(title="Chg")
        sheet2.column_dimensions['A'].width = 10
        sheet2.column_dimensions['B'].width = 20
        sheet2.column_dimensions['C'].width = 2

        sheet3 = book.create_sheet(title="Spread")
        sheet3.column_dimensions['A'].width = 10
        sheet3.column_dimensions['B'].width = 20
        sheet3.column_dimensions['C'].width = 2
        # 저장
        sheet1.append(header1 + remove_dup_data_at_list(headList) + header2 + remove_dup_data_at_list(headList))
        sheet1.append(header2 + titleList + header2 + titleList)
        sheet1.cell(row=3, column=len(header2)+1).value = '매입가(원)'
        sheet1.cell(row=3, column=len(remove_dup_data_at_list(headList)) + 7).value = '판매가(원)' # 6+1
        sheet1.append(datas + buyList + header2 + sellList)

        sheet2.append(header1 + remove_dup_data_at_list(headList) + header2 + remove_dup_data_at_list(headList))
        sheet2.append(header2 + titleList + header2 + titleList)
        sheet2.cell(row=3, column=len(header2) + 1).value = '매입가 Chg.'
        sheet2.cell(row=3, column=len(remove_dup_data_at_list(headList)) + 7).value = '판매가 Chg.'  # 6+1
        sheet2.append(datas + buyPerList + header2 + sellPerList)


        sheet3.append(header1 + remove_dup_data_at_list(headList))
        sheet3.append(header2 + titleList)
        sheet3.cell(row=3, column=len(header2) + 1).value = 'Spread (판매가-매입가)'
        sheet3.append(datas + spreadList)

        book.save(FILENAME)
